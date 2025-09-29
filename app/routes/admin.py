from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import login_required, current_user
from functools import wraps
from app.models import User, Supplier, Shop, Category, Product, Request, get_db, log_action
import csv
import io
from openpyxl import Workbook
from typing import Any, Dict, List, Optional, Union
from werkzeug.wrappers import Response

admin_bp = Blueprint('admin', __name__)

def admin_required(f: Any) -> Any:
    @wraps(f)
    def decorated_function(*args: Any, **kwargs: Any) -> Any:
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Доступ запрещен', 'error')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard() -> str:
    db = get_db()
    
    # Статистика
    stats = {
        'users_count': db.execute('SELECT COUNT(*) FROM users').fetchone()[0],
        'suppliers_count': db.execute('SELECT COUNT(*) FROM suppliers').fetchone()[0],
        'shops_count': db.execute('SELECT COUNT(*) FROM shops').fetchone()[0],
        'products_count': db.execute('SELECT COUNT(*) FROM products').fetchone()[0],
        'orders_count': db.execute('SELECT COUNT(*) FROM orders').fetchone()[0],
        'requests_count': db.execute('SELECT COUNT(*) FROM requests WHERE status = "pending"').fetchone()[0]
    }
    
    return render_template('admin/dashboard.html', stats=stats)

@admin_bp.route('/users')
@login_required
@admin_required
def users():
    users = User.get_all()
    return render_template('admin/users.html', users=users)

@admin_bp.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']
        
        # Проверка существования пользователя
        if User.get_by_email(email):
            flash('Пользователь с таким email уже существует', 'error')
            return render_template('admin/add_user.html')
        
        user_id = User.create(email, password, role)
        
        # Если создается Торговый, создаем запись в таблице suppliers
        if role == 'supplier' and user_id is not None:
            supplier_name = request.form.get('supplier_name', email)
            supplier_info = request.form.get('supplier_info', '')
            Supplier.create(user_id, supplier_name, supplier_info)
        
        log_action(current_user.id, 'create', 'user', user_id)
        flash('Пользователь успешно создан', 'success')
        return redirect(url_for('admin.users'))
    
    return render_template('admin/add_user.html')

@admin_bp.route('/suppliers')
@login_required
@admin_required
def suppliers():
    suppliers = Supplier.get_all()
    return render_template('admin/suppliers.html', suppliers=suppliers)

@admin_bp.route('/suppliers/<int:supplier_id>')
@login_required
@admin_required
def supplier_detail(supplier_id: int):
    db = get_db()
    
    # Получаем информацию о Торговыйе с данными пользователя
    supplier_data = db.execute('''
        SELECT s.*, u.email, u.created_at as user_created_at, u.role
        FROM suppliers s
        JOIN users u ON s.user_id = u.id
        WHERE s.id = ?
    ''', (supplier_id,)).fetchone()
    
    if not supplier_data:
        flash('Торговый не найден', 'error')
        return redirect(url_for('admin.suppliers'))
    
    # Получаем магазины Торговыйа
    shops = Shop.get_by_supplier_id(supplier_id)
    
    # Получаем статистику
    stats = {
        'shops_count': len(shops),
        'pending_requests': db.execute(
            'SELECT COUNT(*) FROM requests WHERE supplier_id = ? AND status = ?',
            (supplier_id, 'pending')
        ).fetchone()[0],
        'total_requests': db.execute(
            'SELECT COUNT(*) FROM requests WHERE supplier_id = ?',
            (supplier_id,)
        ).fetchone()[0]
    }
    
    return render_template('admin/supplier_detail.html', 
                         supplier=supplier_data, shops=shops, stats=stats)

@admin_bp.route('/suppliers/<int:supplier_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_supplier(supplier_id: int):
    db = get_db()
    
    # Получаем информацию о Торговыйе с данными пользователя
    supplier_data = db.execute('''
        SELECT s.*, u.email, u.id as user_id
        FROM suppliers s
        JOIN users u ON s.user_id = u.id
        WHERE s.id = ?
    ''', (supplier_id,)).fetchone()
    
    if not supplier_data:
        flash('Торговый не найден', 'error')
        return redirect(url_for('admin.suppliers'))
    
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        info = request.form.get('info', '')
        
        # Проверяем уникальность email (исключая текущего пользователя)
        existing_user = db.execute(
            'SELECT id FROM users WHERE email = ? AND id != ?',
            (email, supplier_data['user_id'])
        ).fetchone()
        
        if existing_user:
            flash('Пользователь с таким email уже существует', 'error')
            return render_template('admin/edit_supplier.html', supplier=supplier_data)
        
        # Обновляем данные Торговыйа
        db.execute(
            'UPDATE suppliers SET name = ?, info = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
            (name, info, supplier_id)
        )
        
        # Обновляем email пользователя
        db.execute(
            'UPDATE users SET email = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
            (email, supplier_data['user_id'])
        )
        
        db.commit()
        
        log_action(current_user.id, 'update', 'supplier', supplier_id)
        flash('Данные Торговыйа успешно обновлены', 'success')
        return redirect(url_for('admin.supplier_detail', supplier_id=supplier_id))
    
    return render_template('admin/edit_supplier.html', supplier=supplier_data)

@admin_bp.route('/shops')
@login_required
@admin_required
def shops():
    shops = Shop.get_all()
    return render_template('admin/shops.html', shops=shops)

@admin_bp.route('/shops/<int:shop_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_shop(shop_id: int):
    shop = Shop.get_by_id(shop_id)
    
    if not shop:
        flash('Магазин не найден', 'error')
        return redirect(url_for('admin.shops'))
    
    if request.method == 'POST':
        name = request.form['name']
        info = request.form.get('info', '')
        
        Shop.update(shop_id, name, info)
        log_action(current_user.id, 'update', 'shop', shop_id)
        flash('Магазин успешно обновлен', 'success')
        return redirect(url_for('admin.shops'))
    
    return render_template('admin/edit_shop.html', shop=shop)

@admin_bp.route('/shops/<int:shop_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_shop(shop_id: int):
    db = get_db()
    
    # Проверяем, есть ли связанные заявки
    requests_count = db.execute('SELECT COUNT(*) FROM requests WHERE shop_id = ?', (shop_id,)).fetchone()[0]
    
    if requests_count > 0:
        flash('Нельзя удалить магазин, у которого есть заявки', 'error')
        return redirect(url_for('admin.shops'))
    
    # Удаляем магазин
    Shop.delete(shop_id)
    log_action(current_user.id, 'delete', 'shop', shop_id)
    flash('Магазин успешно удален', 'success')
    return redirect(url_for('admin.shops'))

@admin_bp.route('/categories')
@login_required
@admin_required
def categories():
    categories = Category.get_all()
    return render_template('admin/categories.html', categories=categories)

@admin_bp.route('/categories/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_category():
    if request.method == 'POST':
        name = request.form['name']
        description = request.form.get('description', '')
        
        category_id = Category.create(name, description)
        log_action(current_user.id, 'create', 'category', category_id)
        flash('Категория успешно создана', 'success')
        return redirect(url_for('admin.categories'))
    
    return render_template('admin/add_category.html')

@admin_bp.route('/products')
@login_required
@admin_required
def products():
    products = Product.get_all()
    return render_template('admin/products.html', products=products)

@admin_bp.route('/products/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_product():
    if request.method == 'POST':
        name = request.form['name']
        description = request.form.get('description', '')
        price = float(request.form['price'])
        wholesale_price = request.form.get('wholesale_price')
        wholesale_price = float(wholesale_price) if wholesale_price else None
        category_id = request.form.get('category_id')
        category_id = int(category_id) if category_id else None
        
        # Обработка загруженного изображения
        image_url = None
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename and file.filename != '':
                # Проверяем расширение файла
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
                if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                    import os
                    import uuid
                    from werkzeug.utils import secure_filename
                    from flask import current_app
                    
                    # Генерируем уникальное имя файла
                    filename = secure_filename(file.filename)
                    unique_filename = f"{uuid.uuid4().hex}_{filename}"
                    
                    # Создаем путь для сохранения
                    upload_folder = os.path.join(current_app.static_folder, 'uploads')
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    file_path = os.path.join(upload_folder, unique_filename)
                    file.save(file_path)
                    
                    # Сохраняем относительный путь для URL
                    image_url = f"/static/uploads/{unique_filename}"
        
        product_id = Product.create(
            category_id, name, description, 
            price, wholesale_price, image_url
        )
        
        log_action(current_user.id, 'create', 'product', product_id)
        flash('Товар успешно добавлен и доступен во всех магазинах', 'success')
        return redirect(url_for('admin.products'))
    
    categories = Category.get_all()
    return render_template('admin/add_product.html', categories=categories)

@admin_bp.route('/products/<int:product_id>')
@login_required
@admin_required
def product_detail(product_id):
    """Детальный просмотр товара"""
    product = Product.get_by_id(product_id)
    if not product:
        flash('Товар не найден', 'error')
        return redirect(url_for('admin.products'))
    
    # Получить статистику по товару
    db = get_db()
    
    # Количество заявок с этим товаром
    requests_count = db.execute(
        '''SELECT COUNT(DISTINCT r.id) 
           FROM requests r
           JOIN request_items ri ON r.id = ri.request_id
           WHERE ri.product_id = ?''',
        (product_id,)
    ).fetchone()[0]
    
    # Общее количество запрошенного товара
    total_quantity = db.execute(
        '''SELECT COALESCE(SUM(ri.quantity), 0) 
           FROM request_items ri
           WHERE ri.product_id = ?''',
        (product_id,)
    ).fetchone()[0]
    
    # Последние заявки с этим товаром
    recent_requests = db.execute(
        '''SELECT r.id, r.status, r.created_at, s.name as supplier_name,
                  ri.quantity, shop.name as shop_name
           FROM requests r
           JOIN request_items ri ON r.id = ri.request_id
           JOIN shops shop ON r.shop_id = shop.id
           JOIN suppliers s ON shop.supplier_id = s.id
           WHERE ri.product_id = ?
           ORDER BY r.created_at DESC
           LIMIT 10''',
        (product_id,)
    ).fetchall()
    
    stats = {
        'requests_count': requests_count,
        'total_quantity': total_quantity,
        'recent_requests': recent_requests
    }
    
    return render_template('admin/product_detail.html', product=product, stats=stats)

@admin_bp.route('/products/<int:product_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_product(product_id):
    """Редактирование товара"""
    product = Product.get_by_id(product_id)
    if not product:
        flash('Товар не найден', 'error')
        return redirect(url_for('admin.products'))
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        price = request.form.get('price')
        wholesale_price = request.form.get('wholesale_price')
        category_id = request.form.get('category_id')
        
        # Валидация
        if not name:
            flash('Название товара обязательно', 'error')
            return render_template('admin/edit_product.html', product=product, categories=Category.get_all())
        
        if not price:
            flash('Цена обязательна', 'error')
            return render_template('admin/edit_product.html', product=product, categories=Category.get_all())
        
        try:
            price = float(price)
            if price <= 0:
                flash('Цена должна быть больше 0', 'error')
                return render_template('admin/edit_product.html', product=product, categories=Category.get_all())
        except ValueError:
            flash('Неверный формат цены', 'error')
            return render_template('admin/edit_product.html', product=product, categories=Category.get_all())
        
        # Валидация оптовой цены
        wholesale_price_value = None
        if wholesale_price:
            try:
                wholesale_price_value = float(wholesale_price)
                if wholesale_price_value <= 0:
                    flash('Оптовая цена должна быть больше 0', 'error')
                    return render_template('admin/edit_product.html', product=product, categories=Category.get_all())
            except ValueError:
                flash('Неверный формат оптовой цены', 'error')
                return render_template('admin/edit_product.html', product=product, categories=Category.get_all())
        
        # Валидация категории
        category_id_value = None
        if category_id:
            try:
                category_id_value = int(category_id)
            except ValueError:
                flash('Неверный формат категории', 'error')
                return render_template('admin/edit_product.html', product=product, categories=Category.get_all())
        
        # Обработка изображения (используем существующее, если новое не загружено)
        image_url = product['image_url']
        
        # Обновление товара
        Product.update(
            product_id, category_id_value, name, description, 
            price, wholesale_price_value, image_url
        )
        
        log_action(current_user.id, 'update', 'product', product_id)
        flash('Товар успешно обновлен', 'success')
        return redirect(url_for('admin.product_detail', product_id=product_id))
    
    categories = Category.get_all()
    return render_template('admin/edit_product.html', product=product, categories=categories)

@admin_bp.route('/products/<int:product_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_product(product_id):
    product = Product.get_by_id(product_id)
    if not product:
        flash('Товар не найден', 'error')
        return redirect(url_for('admin.products'))
    Product.delete(product_id)
    log_action(current_user.id, 'delete', 'product', product_id)
    flash('Товар успешно удален', 'success')
    return redirect(url_for('admin.products'))

@admin_bp.route('/requests')
@login_required
@admin_required
def requests():
    db = get_db()
    requests = db.execute('''
        SELECT r.*, sh.name as shop_name, s.name as supplier_name,
               COUNT(ri.id) as items_count
        FROM requests r
        JOIN shops sh ON r.shop_id = sh.id
        JOIN suppliers s ON r.supplier_id = s.id
        LEFT JOIN request_items ri ON r.id = ri.request_id
        GROUP BY r.id
        ORDER BY r.created_at DESC
    ''').fetchall()
    
    return render_template('admin/requests.html', requests=requests)

@admin_bp.route('/requests/<int:request_id>')
@login_required
@admin_required
def request_detail(request_id: int) -> Union[str, Response]:
    db = get_db()
    
    request_info = db.execute('''
        SELECT r.*, sh.name as shop_name, s.name as supplier_name,
               u.email as supplier_email
        FROM requests r
        JOIN shops sh ON r.shop_id = sh.id
        JOIN suppliers s ON r.supplier_id = s.id
        JOIN users u ON s.user_id = u.id
        WHERE r.id = ?
    ''', (request_id,)).fetchone()
    
    if not request_info:
        flash('Заявка не найдена', 'error')
        return redirect(url_for('admin.requests'))
    
    # Получаем детальную информацию о товарах
    items = db.execute('''
        SELECT ri.*, 
               p.name as product_name, 
               p.price, 
               p.wholesale_price,
               p.description as product_description,
               c.name as category_name
        FROM request_items ri
        JOIN products p ON ri.product_id = p.id
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE ri.request_id = ?
        ORDER BY p.name
    ''', (request_id,)).fetchall()
    
    # Вычисляем итоги
    total_cost = 0
    total_quantity = 0
    total_retail_cost = 0
    total_wholesale_cost = 0
    
    for item in items:
        item_total = item['price'] * item['quantity']
        total_cost += item_total
        total_quantity += item['quantity']
        
        # Розничная стоимость
        total_retail_cost += item['price'] * item['quantity']
        
        # Оптовая стоимость
        wholesale_price = item['wholesale_price'] if item['wholesale_price'] else item['price'] * 0.85
        total_wholesale_cost += wholesale_price * item['quantity']
    
    # Дополнительная статистика
    stats = {
        'total_cost': total_cost,
        'total_quantity': total_quantity,
        'total_retail_cost': total_retail_cost,
        'total_wholesale_cost': total_wholesale_cost,
        'items_count': len(items),
        'avg_price_per_item': total_cost / len(items) if len(items) > 0 else 0,
        'avg_price_per_unit': total_cost / total_quantity if total_quantity > 0 else 0
    }
    
    return render_template('admin/request_detail.html', request=request_info, items=items, stats=stats)

@admin_bp.route('/requests/<int:request_id>/export')
@login_required
@admin_required
def export_request(request_id: int) -> Union[str, Response]:
    """Экспорт заявки в Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import io

    db = get_db()
    
    # Получаем информацию о заявке
    request_info = db.execute('''
        SELECT r.*, sh.name as shop_name, sh.business_type, s.name as supplier_name
        FROM requests r
        JOIN shops sh ON r.shop_id = sh.id
        JOIN suppliers s ON r.supplier_id = s.id
        WHERE r.id = ?
    ''', (request_id,)).fetchone()
    
    if not request_info:
        flash('Заявка не найдена', 'error')
        return redirect(url_for('admin.requests'))
    
    # Получаем товары
    items = db.execute('''
        SELECT ri.*, 
               p.name as product_name, 
               p.price, 
               p.wholesale_price,
               p.description as product_description,
               c.name as category_name
        FROM request_items ri
        JOIN products p ON ri.product_id = p.id
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE ri.request_id = ?
        ORDER BY p.name
    ''', (request_id,)).fetchall()
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = f"Заявка_{request_id}"
    
    # Стили
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Заголовок документа
    ws.merge_cells('A1:E1')
    ws['A1'] = f"ЗАЯВКА #{request_id}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    # Информация о заявке
    row = 3
    ws[f'A{row}'] = "Информация о заявке"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    info_data = [
        ("Магазин:", request_info['shop_name']),
        ("Тип организации:", request_info['business_type'] or 'ИП'),
        ("Дата отправки:", request_info['created_at']),
    ]
    
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Пустая строка
    row += 2
    
    # Заголовок таблицы товаров
    ws[f'A{row}'] = "Товары в заявке"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    # Заголовки таблицы
    headers = ['Товар', 'Количество', 'Цена за ед.', 'Сумма', '% от общей суммы']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    row += 1
    
    # Данные товаров
    total_cost = 0
    total_quantity = 0
    
    for item in items:
        item_total = item['price'] * item['quantity']
        total_cost += item_total
        total_quantity += item['quantity']
    
    for item in items:
        item_total = item['price'] * item['quantity']
        percentage = (item_total / total_cost * 100) if total_cost > 0 else 0
        
        data = [
            item['product_name'],
            f"{item['quantity']} шт.",
            f"{item['price']:.0f} ₸",
            f"{item_total:.0f} ₸",
            f"{percentage:.1f}%"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [3, 4]:  # Цена и сумма
                cell.alignment = Alignment(horizontal='right')
            elif col in [2, 5]:  # Количество и процент
                cell.alignment = center_alignment
        
        row += 1
    
    # Итоговая строка
    ws[f'A{row}'] = "ИТОГО:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"{total_quantity} шт."
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].alignment = center_alignment
    ws[f'D{row}'] = f"{total_cost:.0f} ₸"
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'D{row}'].alignment = Alignment(horizontal='right')
    ws[f'E{row}'] = "100%"
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'E{row}'].alignment = center_alignment
    
    # Применяем границы к итоговой строке
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    
    # Автоподбор ширины колонок
    for col in range(1, 6):
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    # Сохраняем в память
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Создаем ответ
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    # Используем ASCII имя файла для избежания проблем с кодировкой
    filename = f'Request_{request_id}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@admin_bp.route('/requests/<int:request_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_request(request_id):
    """Удаление заявки"""
    # Проверяем существование заявки
    db = get_db()
    request_info = db.execute('SELECT * FROM requests WHERE id = ?', (request_id,)).fetchone()
    if not request_info:
        flash('Заявка не найдена', 'error')
        return redirect(url_for('admin.requests'))
    
    # Удаляем заявку (метод delete удалит и связанные позиции)
    Request.delete(request_id)
    log_action(current_user.id, 'delete', 'request', request_id)
    flash('Заявка успешно удалена', 'success')
    return redirect(url_for('admin.requests'))

@admin_bp.route('/requests/<int:request_id>/mark_processed', methods=['POST'])
@login_required
@admin_required
def mark_request_processed(request_id: int) -> Response:
    db = get_db()
    db.execute('UPDATE requests SET status = "completed" WHERE id = ?', (request_id,))
    db.commit()
    
    log_action(current_user.id, 'update', 'request', request_id)
    flash('Заявка отмечена как обработанная', 'success')
    return redirect(url_for('admin.request_detail', request_id=request_id))

@admin_bp.route('/requests/<int:request_id>/reopen', methods=['POST'])
@login_required
@admin_required
def reopen_request(request_id: int) -> Response:
    db = get_db()
    db.execute('UPDATE requests SET status = "pending" WHERE id = ?', (request_id,))
    db.commit()
    
    log_action(current_user.id, 'update', 'request', request_id)
    flash('Заявка возвращена в обработку для редактирования', 'success')
    return redirect(url_for('admin.request_detail', request_id=request_id))

@admin_bp.route('/reports')
@login_required
@admin_required
def reports() -> str:
    return render_template('admin/reports.html')

@admin_bp.route('/reports/export/<report_type>')
@login_required
@admin_required
def export_report(report_type: str) -> Response:
    db = get_db()
    
    if report_type == 'products':
        data = db.execute('''
            SELECT p.name, p.price, p.wholesale_price,
                   c.name as category
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            ORDER BY p.name
        ''').fetchall()
        
        filename = 'products_report.xlsx'
        headers = ['Название', 'Цена', 'Опт. цена', 'Категория']
    
    elif report_type == 'shops':
        data = db.execute('''
            SELECT sh.name, s.name as supplier_name, sh.created_at
            FROM shops sh
            LEFT JOIN suppliers s ON sh.supplier_id = s.id
            ORDER BY sh.name
        ''').fetchall()
        
        filename = 'shops_report.xlsx'
        headers = ['Магазин', 'Торговый', 'Создан']
    
    else:
        flash('Неизвестный тип отчета', 'error')
        return redirect(url_for('admin.reports'))
    
    # Создание Excel файла
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.title()
    
    # Заголовки
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Данные
    for row, item in enumerate(data, 2):
        for col, value in enumerate(item, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Создание ответа
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response